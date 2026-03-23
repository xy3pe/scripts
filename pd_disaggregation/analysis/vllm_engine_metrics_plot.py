"""
vLLM Engine 指标解析与绘图
从 prefill.log / decode.log（或 1P多D 时的 decode_0.log、decode_1.log、…）中解析 vLLM Engine 指标
（throughput、Running/Waiting、KV cache 等），输出 Excel：原始数据 + 折线图。

单目录模式：输入目录下直接含 prefill.log 与 decode.log（或 decode_0.log、decode_1.log 等）时，
在该目录生成 vllm_engine_metrics_analysis.xlsx
（原始数据表 + Throughput / Running_Waiting / Cache 三类图，每类「原始 + 采样」两张图并排）。

批量模式：输入目录下无日志、仅有子目录（如 batch_50, batch_60）时，每个子目录生成各自的
vllm_engine_metrics_analysis.xlsx，并在父目录生成 vllm_engine_metrics_sweep.xlsx（7 指标 × prefill+decode，
每 sheet 横轴 index、多子目录折线对比）。

用法：
  python vllm_engine_metrics_plot.py <log_dir> [--output 文件名.xlsx]
  # --output 仅指定文件名，输出路径始终在 prefill.log 同级；批量时汇总表固定为 vllm_engine_metrics_sweep.xlsx
"""

import re
import os
import glob
import argparse
import colorsys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


# ── 配置（可在此集中修改）────────────────────────────────────────────────────
# 解析
ENGINE_LINE_PATTERN = re.compile(
    r"(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})"
    r".*?"
    r"Engine\s+\d+:"
    r"\s*Avg prompt throughput:\s*([\d.]+)\s*tokens/s,"
    r"\s*Avg generation throughput:\s*([\d.]+)\s*tokens/s,"
    r"\s*Running:\s*(\d+)\s*reqs,"
    r"\s*Waiting:\s*(\d+)\s*reqs,"
    r"\s*GPU KV cache usage:\s*([\d.]+)%,"
    r"\s*Prefix cache hit rate:\s*([\d.]+)%,"
    r"\s*External prefix cache hit rate:\s*([\d.]+)%",
    re.IGNORECASE,
)
COLUMNS = [
    "avg_prompt_throughput",
    "avg_generation_throughput",
    "running",
    "waiting",
    "gpu_kv_cache_usage_pct",
    "prefix_cache_hit_rate_pct",
    "external_prefix_cache_hit_rate_pct",
]
FORWARD_FILL_COLUMNS = {"gpu_kv_cache_usage_pct", "prefix_cache_hit_rate_pct", "external_prefix_cache_hit_rate_pct"}
TIME_ALIGN_TOLERANCE_SEC = 5  # 时间对齐容差（秒），该范围内 prefill/decode 视为同一时刻

# Excel 样式与列宽
FONT_NAME = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
BORDER = Border(
    left=Side(border_style="thin", color="BFBFBF"),
    right=Side(border_style="thin", color="BFBFBF"),
    top=Side(border_style="thin", color="BFBFBF"),
    bottom=Side(border_style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center")
COL_WIDTH_TIME = 20
COL_WIDTH_INDEX = 10
COL_WIDTH_DATA = 16
RAW_HEADERS = ["index", "timestamp"] + COLUMNS

# 图 sheet 布局与采样
CHART_SHEET_DATA_START_ROW = 55
CHART2_ANCHOR = "J1"
MAX_CHART_POINTS = 100
LINE_WIDTH_EMU = 28000
# 折线图尺寸（宽×高，单位与 Excel 列宽一致）
CHART_WIDTH = 24
CHART_HEIGHT = 16
CHART_SWEEP_WIDTH = 24
CHART_SWEEP_HEIGHT = 16

# 汇总表（批量场景）指标显示名
METRIC_SWEEP_TITLES = {
    "avg_prompt_throughput": "prompt throughput",
    "avg_generation_throughput": "generation throughput",
    "running": "Running",
    "waiting": "Waiting",
    "gpu_kv_cache_usage_pct": "KV cache usage",
    "prefix_cache_hit_rate_pct": "Prefix hit",
    "external_prefix_cache_hit_rate_pct": "External prefix hit",
}


def parse_engine_line(line):
    """从一行日志中解析出指标字典，不匹配则返回 None。"""
    m = ENGINE_LINE_PATTERN.search(line)
    if not m:
        return None
    return {
        "timestamp": m.group(1).strip(),
        "avg_prompt_throughput": float(m.group(2)),
        "avg_generation_throughput": float(m.group(3)),
        "running": int(m.group(4)),
        "waiting": int(m.group(5)),
        "gpu_kv_cache_usage_pct": float(m.group(6)),
        "prefix_cache_hit_rate_pct": float(m.group(7)),
        "external_prefix_cache_hit_rate_pct": float(m.group(8)),
    }


def load_log_series(log_path):
    """从单个日志文件中按行解析所有 Engine 指标，返回列表 of dict。"""
    rows = []
    if not os.path.isfile(log_path):
        return rows
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            d = parse_engine_line(line)
            if d is not None:
                rows.append(d)
    return rows


def _decode_multi_log_index(file_path):
    """从 decode_3.log 提取 3，用于多 Decode 文件排序；非 decode_N.log 返回 -1。"""
    basename = os.path.basename(file_path)
    m = re.match(r"decode_(\d+)\.log", basename)
    return int(m.group(1)) if m else -1


def load_prefill_decode(log_dir):
    """
    从 log_dir 下读取 prefill 与 decode 日志，返回 (prefill_rows, decode_rows_list)。
    - 1P1D：存在 decode.log 时只读该文件，返回 decode_rows_list = [decode_rows]。
    - 1P多D：不存在 decode.log 时，枚举 decode_0.log、decode_1.log、… 按编号排序后分别加载，
      每个文件单独成一份，返回 decode_rows_list = [decode_0_rows, decode_1_rows, ...]。
    不同 decode 视为不同数据源，不做合并。
    """
    log_dir = os.path.abspath(log_dir)
    prefill_path = os.path.join(log_dir, "prefill.log")
    decode_single = os.path.join(log_dir, "decode.log")
    prefill_rows = load_log_series(prefill_path)
    if os.path.isfile(decode_single):
        decode_rows_list = [load_log_series(decode_single)]
    else:
        pattern = os.path.join(log_dir, "decode_*.log")
        decode_files = sorted(glob.glob(pattern), key=_decode_multi_log_index)
        decode_rows_list = [load_log_series(path) for path in decode_files]
    return prefill_rows, decode_rows_list


def _dir_has_logs(path):
    """目录下是否直接包含 prefill.log、decode.log 或 decode_*.log（视为单目录模式）。"""
    p = os.path.join(path, "prefill.log")
    d = os.path.join(path, "decode.log")
    if os.path.isfile(p) or os.path.isfile(d):
        return True
    return len(glob.glob(os.path.join(path, "decode_*.log"))) > 0


def _parse_timestamp(ts_str):
    """将 'MM-DD HH:MM:SS' 转为可比较的 float（固定年份保证顺序）。"""
    try:
        dt = datetime.strptime(ts_str.strip(), "%m-%d %H:%M:%S")
        return dt.replace(year=2024).timestamp()
    except (ValueError, TypeError):
        return None


def align_by_time(prefill_rows, decode_rows_list):
    """
    按时间对齐 prefill 与多路 decode：将时间差在 TIME_ALIGN_TOLERANCE_SEC 内的样本视为同一时刻合并为一行。
    返回 [(time_label, p_row, d_row_0, d_row_1, ...), ...]，按时间排序；decode_rows_list 为 [decode_0_rows, decode_1_rows, ...]。
    """
    n_decode = len(decode_rows_list)

    def with_ts(rows, src, decode_idx=0):
        out = []
        for r in rows:
            t = _parse_timestamp(r.get("timestamp", ""))
            if t is not None:
                out.append((t, src, decode_idx, r))
        return out

    events = with_ts(prefill_rows, "p")
    for di, rows in enumerate(decode_rows_list):
        events.extend(with_ts(rows, "d", di))
    if not events:
        return []
    events.sort(key=lambda x: x[0])

    clusters = []
    i = 0
    while i < len(events):
        t0 = events[i][0]
        cluster = [events[i]]
        i += 1
        while i < len(events) and events[i][0] - t0 <= TIME_ALIGN_TOLERANCE_SEC:
            cluster.append(events[i])
            i += 1
        clusters.append(cluster)

    result = []
    for cluster in clusters:
        ts = [x[0] for x in cluster]
        rep_t = sum(ts) / len(ts)
        label = datetime.fromtimestamp(rep_t).strftime("%m-%d %H:%M:%S")
        p_entries = [(x[0], x[3]) for x in cluster if x[1] == "p"]
        p_row = min(p_entries, key=lambda e: abs(e[0] - rep_t))[1] if p_entries else None
        d_rows = [None] * n_decode
        for di in range(n_decode):
            d_entries = [(x[0], x[3]) for x in cluster if x[1] == "d" and x[2] == di]
            d_rows[di] = min(d_entries, key=lambda e: abs(e[0] - rep_t))[1] if d_entries else None
        result.append((label, p_row, *d_rows))
    return result


def _set_sheet_column_widths(ws, widths):
    """按列表依次设置各列宽。widths[i] 为第 i+1 列宽度。"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _max_len(prefill_rows, decode_rows_list):
    """prefill 与所有 decode 列表的最大长度。"""
    decode_lens = [len(dr) for dr in decode_rows_list] if decode_rows_list else [0]
    return max(len(prefill_rows), max(decode_lens, default=0), 1)


def _pad(rows, length, fill=None):
    """将 rows 填充到 length 长度，不足用 fill。"""
    return list(rows) + [fill] * (length - len(rows))


def write_raw_sheet(wb, prefill_rows, decode_rows_list):
    """写入原始数据：一个 Prefill sheet + 每个 decode 一份 sheet（Decode 原始数据 或 Decode_0/Decode_1 原始数据）。"""
    # Prefill
    ws_prefill = wb.create_sheet("Prefill 原始数据", 0)
    for c, h in enumerate(RAW_HEADERS, 1):
        cell = ws_prefill.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for r, row in enumerate(prefill_rows, 2):
        for col_idx, val in enumerate([r - 2, row.get("timestamp", "")] + [row.get(key) for key in COLUMNS], 1):
            cell = ws_prefill.cell(row=r, column=col_idx, value=val)
            cell.alignment = CENTER
    _set_sheet_column_widths(ws_prefill, [COL_WIDTH_INDEX, COL_WIDTH_TIME] + [COL_WIDTH_DATA] * len(COLUMNS))
    # Decode：每个 decode 单独一个 sheet
    for di, decode_rows in enumerate(decode_rows_list):
        sheet_name = "Decode 原始数据" if len(decode_rows_list) == 1 else f"Decode_{di} 原始数据"
        ws_decode = wb.create_sheet(_sanitize_sheet_title(sheet_name), 1 + di)
        for c, h in enumerate(RAW_HEADERS, 1):
            cell = ws_decode.cell(row=1, column=c, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        for r, row in enumerate(decode_rows, 2):
            for col_idx, val in enumerate([r - 2, row.get("timestamp", "")] + [row.get(key) for key in COLUMNS], 1):
                cell = ws_decode.cell(row=r, column=col_idx, value=val)
                cell.alignment = CENTER
        _set_sheet_column_widths(ws_decode, [COL_WIDTH_INDEX, COL_WIDTH_TIME] + [COL_WIDTH_DATA] * len(COLUMNS))


def _downsample_indices(n, max_pts):
    """
    返回用于绘图的索引列表。若 n <= max_pts 则全选；否则均匀采样，保留首尾，共约 max_pts 个点。
    """
    if n <= max_pts:
        return list(range(n))
    step = (n - 1) / (max_pts - 1)
    indices = [0]
    for i in range(1, max_pts - 1):
        indices.append(int(round(i * step)))
    indices.append(n - 1)
    return indices


def _write_data_block_aligned(ws, row0, headers, columns_subset, aligned_rows):
    """
    在 sheet 的 row0 起写入一块表头+数据；
    aligned_rows = [(time_label, p_row, d_row_0, d_row_1, ...), ...]，多路 decode 时有多列。
    表头由调用方传入：index, prefill_time, decode_0_time, decode_1_time, ... , prefill_xx, decode_0_xx, decode_1_xx, ...
    仅 FORWARD_FILL_COLUMNS 中的指标在缺侧时做前向填充；其余指标该侧留空。
    """
    n_rows = len(aligned_rows)
    if not n_rows:
        return 0
    n_decode = len(aligned_rows[0]) - 2
    index_col = 1
    prefill_time_col = 2
    time_cols = 1 + n_decode
    data_start_col = 2 + time_cols
    cols_per_metric = 1 + n_decode

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row0, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    last_p = {col: None for col in columns_subset if col in FORWARD_FILL_COLUMNS}
    last_d = [{col: None for col in columns_subset if col in FORWARD_FILL_COLUMNS} for _ in range(n_decode)]
    for ri, row_tuple in enumerate(aligned_rows):
        sheet_row = row0 + 1 + ri
        p_row = row_tuple[1]
        d_rows = list(row_tuple[2:2 + n_decode])
        ws.cell(row=sheet_row, column=index_col, value=ri)
        ws.cell(row=sheet_row, column=index_col).alignment = CENTER
        pre_ts = p_row.get("timestamp") if p_row else None
        ws.cell(row=sheet_row, column=prefill_time_col, value=pre_ts if pre_ts else "")
        ws.cell(row=sheet_row, column=prefill_time_col).alignment = CENTER
        for di, d_row in enumerate(d_rows):
            dec_ts = d_row.get("timestamp") if d_row else None
            ws.cell(row=sheet_row, column=prefill_time_col + 1 + di, value=dec_ts if dec_ts else "")
            ws.cell(row=sheet_row, column=prefill_time_col + 1 + di).alignment = CENTER
        for ci, col in enumerate(columns_subset):
            base_col = data_start_col + ci * cols_per_metric
            raw_p = p_row.get(col) if p_row else None
            if col in FORWARD_FILL_COLUMNS:
                if raw_p is not None:
                    last_p[col] = raw_p
                pv = raw_p if raw_p is not None else last_p.get(col)
            else:
                pv = raw_p
            ws.cell(row=sheet_row, column=base_col, value=pv if pv is not None else "")
            ws.cell(row=sheet_row, column=base_col).alignment = CENTER
            for di, d_row in enumerate(d_rows):
                raw_d = d_row.get(col) if d_row else None
                if col in FORWARD_FILL_COLUMNS:
                    if raw_d is not None:
                        last_d[di][col] = raw_d
                    dv = raw_d if raw_d is not None else last_d[di].get(col)
                else:
                    dv = raw_d
                cell_d = ws.cell(row=sheet_row, column=base_col + 1 + di, value=dv if dv is not None else "")
                cell_d.alignment = CENTER
    return n_rows


def _build_line_chart(ws, row0, n_rows, columns_subset, titles, chart_title, num_metrics, num_decodes, palette):
    """根据已有数据区建折线图；每指标 1 个 prefill 列 + num_decodes 个 decode 列。"""
    time_cols = 1 + num_decodes
    data_start_col = 2 + time_cols
    cols_per_metric = 1 + num_decodes
    chart = LineChart()
    chart.title = chart_title
    chart.style = 10
    chart.y_axis.title = titles[0] if len(titles) == 1 else "Value"
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    cats = Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + n_rows)
    chart.set_categories(cats)
    decode_labels = ["decode"] if num_decodes == 1 else [f"decode_{i}" for i in range(num_decodes)]
    for ci in range(num_metrics):
        base_col = data_start_col + ci * cols_per_metric
        ref_p = Reference(ws, min_col=base_col, min_row=row0, max_row=row0 + n_rows)
        chart.add_data(ref_p, titles_from_data=True)
        chart.series[-1].tx = SeriesLabel(v=f"prefill {titles[ci]}")
        chart.series[-1].smooth = True
        for di in range(num_decodes):
            ref_d = Reference(ws, min_col=base_col + 1 + di, min_row=row0, max_row=row0 + n_rows)
            chart.add_data(ref_d, titles_from_data=True)
            chart.series[-1].tx = SeriesLabel(v=f"{decode_labels[di]} {titles[ci]}")
            chart.series[-1].smooth = True
    _style_series_multi(chart, num_metrics, num_decodes, palette)
    return chart


def _chart_data_sheet(wb, prefill_rows, decode_rows_list, name, columns_subset, titles):
    """
    一个 sheet 两张图：按时间对齐 prefill 与多路 decode，相近时刻的数据在同一行对比。
    原始数据图 + 采样图；表头含 prefill_time、decode_0_time/decode_1_time/... 及对应指标列。
    """
    n_decode = len(decode_rows_list)
    aligned = align_by_time(prefill_rows, decode_rows_list)
    if not aligned:
        n = _max_len(prefill_rows, decode_rows_list)
        pad = [None] * n_decode
        aligned = []
        for i in range(n):
            p_row = prefill_rows[i] if i < len(prefill_rows) else None
            d_rows = [decode_rows_list[di][i] if i < len(decode_rows_list[di]) else None for di in range(n_decode)]
            aligned.append((str(i), p_row, *d_rows))
    n = len(aligned)
    indices_sampled = _downsample_indices(n, MAX_CHART_POINTS)
    aligned_sampled = [aligned[i] for i in indices_sampled]
    n_sampled = len(aligned_sampled)

    time_headers = ["index", "prefill_time"]
    time_headers += ["decode_time"] if n_decode == 1 else [f"decode_{i}_time" for i in range(n_decode)]
    headers = list(time_headers)
    for col in columns_subset:
        headers.append(f"prefill_{col}")
        if n_decode == 1:
            headers.append(f"decode_{col}")
        else:
            for di in range(n_decode):
                headers.append(f"decode_{di}_{col}")

    ws = wb.create_sheet(name)
    num_metrics = len(columns_subset)
    palette = _chart_palette(num_metrics, n_decode)

    row0_full = CHART_SHEET_DATA_START_ROW
    _write_data_block_aligned(ws, row0_full, headers, columns_subset, aligned)
    n_full = n
    row0_sampled = row0_full + n_full + 2
    _write_data_block_aligned(ws, row0_sampled, headers, columns_subset, aligned_sampled)

    chart1 = _build_line_chart(ws, row0_full, n_full, columns_subset, titles, f"{name} (原始数据)", num_metrics, n_decode, palette)
    ws.add_chart(chart1, "A1")
    chart2 = _build_line_chart(ws, row0_sampled, n_sampled, columns_subset, titles, f"{name} (采样)", num_metrics, n_decode, palette)
    ws.add_chart(chart2, CHART2_ANCHOR)
    n_cols = len(headers)
    _set_sheet_column_widths(ws, [COL_WIDTH_INDEX] + [COL_WIDTH_TIME] * (1 + n_decode) + [COL_WIDTH_DATA] * (n_cols - 2 - n_decode))
    return ws


def _chart_palette(num_metrics, num_decodes=1):
    """
    每指标一色系：prefill 深色，decode 用 1～num_decodes 个浅色（由深到浅）。
    返回 list of (prefill_hex, [decode_0_hex, decode_1_hex, ...])。
    """
    # 每行：prefill 深色，decode 浅色列表（支持多路 D，每系 6 个）
    base = [
        ("1B5E9E", ["64B5F6", "90CAF9", "BBDEFB", "E3F2FD", "B3E5FC", "F5F5F5"]),   # 蓝
        ("2E7D32", ["81C784", "A5D6A7", "C8E6C9", "E8F5E9", "DCEDC8", "F1F8E9"]),   # 绿
        ("C62828", ["E57373", "EF9A9A", "FFCDD2", "FFEBEE", "F8BBD9", "FCE4EC"]),   # 红
        ("6A1B9A", ["B39DDB", "CE93D8", "E1BEE7", "F3E5F5", "EDE7F6", "F8BBD9"]),   # 紫
        ("E65100", ["FFB74D", "FFCC80", "FFE0B2", "FFF3E0", "FFECB3", "FFF8E1"]),   # 橙
    ]
    out = []
    for i in range(num_metrics):
        prefill_hex, decode_hexes = base[i % len(base)]
        out.append((prefill_hex, decode_hexes[:num_decodes]))
    return out


def _style_series_multi(chart, num_metrics, num_decodes, colors):
    """
    series 顺序：每指标 1 个 prefill + num_decodes 个 decode。
    同指标同色系：prefill 深色，decode_0/decode_1 用对应浅色。
    """
    series_per_metric = 1 + num_decodes
    for i in range(len(chart.series)):
        metric_idx = i // series_per_metric
        pos = i % series_per_metric
        t = colors[metric_idx] if metric_idx < len(colors) else ("333333", ["999999"])
        prefill_hex, decode_hexes = t
        color = prefill_hex if pos == 0 else (decode_hexes[pos - 1] if pos - 1 < len(decode_hexes) else "999999")
        line = chart.series[i].graphicalProperties.line
        line.solidFill = color
        line.dashStyle = "solid"
        line.width = LINE_WIDTH_EMU


def write_charts(wb, prefill_rows, decode_rows_list):
    """创建三张图的数据表并插入折线图（支持多路 decode）。"""
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows_list,
        name="图1_Throughput",
        columns_subset=["avg_prompt_throughput", "avg_generation_throughput"],
        titles=["Avg prompt throughput (tokens/s)", "Avg generation throughput (tokens/s)"],
    )
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows_list,
        name="图2_Running_Waiting",
        columns_subset=["running", "waiting"],
        titles=["Running (reqs)", "Waiting (reqs)"],
    )
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows_list,
        name="图3_Cache",
        columns_subset=[
            "gpu_kv_cache_usage_pct",
            "prefix_cache_hit_rate_pct",
            "external_prefix_cache_hit_rate_pct",
        ],
        titles=[
            "GPU KV cache usage (%)",
            "Prefix cache hit rate (%)",
            "External prefix cache hit rate (%)",
        ],
    )


def _process_one_dir(log_dir, out_path, prefill_rows=None, decode_rows_list=None):
    """处理单个目录并写入 Excel；返回 True 成功，False 无有效数据。若未传入则从 log_dir 加载。"""
    if prefill_rows is None or decode_rows_list is None:
        prefill_rows, decode_rows_list = load_prefill_decode(log_dir)
    if not prefill_rows and not any(decode_rows_list):
        return False
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_raw_sheet(wb, prefill_rows, decode_rows_list)
    write_charts(wb, prefill_rows, decode_rows_list)
    wb.save(out_path)
    return True


def _mean_over_rows(rows, col):
    """对多行数据取某列均值，忽略 None；无有效值时返回 None。"""
    vals = [r.get(col) for r in rows if r.get(col) is not None]
    if not vals:
        return None
    return sum(vals) / len(vals)


def _sanitize_sheet_title(title):
    """Excel sheet 名不能含 \\ / * ? [ ] : ，替换为下划线并截断至 31 字符。"""
    for c in r'\/*?[]:':
        title = title.replace(c, "_")
    return title[:31] if len(title) > 31 else title


def _write_sweep_data_block(ws, row0, batch_series, batch_entries, indices, timestamps=None):
    """写入一块汇总数据：index + 时间戳（可选）+ 各 batch 列，仅 indices 中的行。"""
    n_batches = len(batch_entries)
    off = 2 if timestamps is not None else 1  # 数据列起始列（index 占 1，时间戳可选占 1）
    n_cols = off + n_batches
    ws.cell(row=row0, column=1, value="index")
    if timestamps is not None:
        ws.cell(row=row0, column=2, value="时间戳")
    for bi, (name, _, _) in enumerate(batch_entries):
        ws.cell(row=row0, column=off + 1 + bi, value=name)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row0, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for ri, idx in enumerate(indices):
        sheet_row = row0 + 1 + ri
        ws.cell(row=sheet_row, column=1, value=idx)
        if timestamps is not None:
            ts = timestamps[idx] if idx < len(timestamps) else ""
            ws.cell(row=sheet_row, column=2, value=ts)
        for bi, series in enumerate(batch_series):
            val = series[idx] if idx < len(series) else None
            ws.cell(row=sheet_row, column=off + 1 + bi, value=val if val is not None else "")
        for c in range(1, n_cols + 1):
            ws.cell(row=sheet_row, column=c).alignment = CENTER


def _add_sweep_chart(ws, row0, n_rows, batch_entries, title, palette, has_timestamp_col=True):
    """在 ws 上根据 row0 起的数据区添加一张汇总折线图。has_timestamp_col 为 True 时第 2 列为时间戳，数据从第 3 列起。"""
    n_batches = len(batch_entries)
    data_start_col = 3 if has_timestamp_col else 2
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = title
    chart.width = CHART_SWEEP_WIDTH
    chart.height = CHART_SWEEP_HEIGHT
    cats = Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + n_rows)
    chart.set_categories(cats)
    for bi in range(n_batches):
        ref = Reference(ws, min_col=data_start_col + bi, min_row=row0, max_row=row0 + n_rows)
        chart.add_data(ref, titles_from_data=True)
        chart.series[-1].tx = SeriesLabel(v=batch_entries[bi][0])
        chart.series[-1].smooth = True
    for i in range(len(chart.series)):
        line = chart.series[i].graphicalProperties.line
        # palette 传入每条线一个颜色（hex），避免多 batch 时颜色重复
        if palette and i < len(palette):
            line.solidFill = palette[i]
        else:
            line.solidFill = "333333"
        line.dashStyle = "solid"
        line.width = LINE_WIDTH_EMU
    return chart


def _series_colors(n: int) -> list[str]:
    """为 n 条折线生成尽量不同的颜色（hex），用于 sweep 多子目录区分。"""
    if n <= 0:
        return []
    # 调低饱和度与明度：避免 Excel 里颜色过于“鲜艳/刺眼”，同时仍保持可区分。
    s = 0.52
    v = 0.78
    colors = []
    for i in range(n):
        hue = i / n
        r, g, b = colorsys.hsv_to_rgb(hue, s, v)
        colors.append(f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}")
    return colors


def build_sweep_excel(log_dir, batch_entries, sweep_filename):
    """
    批量场景：每个指标 × (prefill + decode_0 + decode_1 + ...) 一个 Sheet；每个 sheet 横轴 = index，
    纵轴 = 指标值，多列对比各 batch；含原始数据图 + 采样图。多路 decode 时每个 decode 单独一个 sheet。
    """
    if not batch_entries:
        return
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    max_decodes = max(len(drl) for _, _, drl in batch_entries)
    sources = ["prefill"] + [f"decode_{i}" for i in range(max_decodes)]

    for col in COLUMNS:
        title_base = METRIC_SWEEP_TITLES.get(col, col)
        for si, source in enumerate(sources):
            sheet_title = _sanitize_sheet_title(f"{title_base}_{source}")
            ws = wb.create_sheet(sheet_title, len(wb.sheetnames))
            batch_series = []
            first_batch_rows = None
            for _name, prefill_rows, decode_rows_list in batch_entries:
                if source == "prefill":
                    rows = prefill_rows
                else:
                    di = int(source.split("_")[-1])
                    rows = decode_rows_list[di] if di < len(decode_rows_list) else []
                if first_batch_rows is None:
                    first_batch_rows = rows
                batch_series.append([r.get(col) for r in rows] if rows else [])
            n_points = max(len(s) for s in batch_series) if batch_series else 0
            if n_points == 0:
                continue
            timestamps = [first_batch_rows[i]["timestamp"] if first_batch_rows and i < len(first_batch_rows) else "" for i in range(n_points)]
            indices_full = list(range(n_points))
            indices_sampled = _downsample_indices(n_points, MAX_CHART_POINTS)
            n_sampled = len(indices_sampled)

            row0_full = CHART_SHEET_DATA_START_ROW
            _write_sweep_data_block(ws, row0_full, batch_series, batch_entries, indices_full, timestamps=timestamps)
            row0_sampled = row0_full + n_points + 2
            _write_sweep_data_block(ws, row0_sampled, batch_series, batch_entries, indices_sampled, timestamps=timestamps)

            n_batches = len(batch_entries)
            colors = _series_colors(n_batches)
            chart1 = _add_sweep_chart(ws, row0_full, n_points, batch_entries, f"{title_base} ({source}) 原始数据", colors, has_timestamp_col=True)
            ws.add_chart(chart1, "A1")
            chart2 = _add_sweep_chart(ws, row0_sampled, n_sampled, batch_entries, f"{title_base} ({source}) 采样", colors, has_timestamp_col=True)
            ws.add_chart(chart2, CHART2_ANCHOR)

            _set_sheet_column_widths(ws, [COL_WIDTH_INDEX, COL_WIDTH_TIME] + [COL_WIDTH_DATA] * n_batches)

    out_path = os.path.join(log_dir, sweep_filename)
    wb.save(out_path)
    print(f"已保存汇总表: {out_path}")


def process_log_dir(log_dir, output_filename):
    """
    仅两种情形：
    1）log_dir 下直接有 prefill.log / decode.log：在该目录下生成 Excel，返回 1。
    2）log_dir 下无日志，仅有若干直接子目录且子目录下有 prefill.log / decode.log：对每个此类子目录生成 Excel，汇总表仅写在父目录。
    """
    log_dir = os.path.abspath(log_dir)

    if _dir_has_logs(log_dir):
        # 单目录模式只应生成「分析」表，不应生成 sweep 命名文件（sweep 仅用于批量时的父目录汇总）
        if "sweep" in output_filename.lower():
            output_filename = "vllm_engine_metrics_analysis.xlsx"
        out_path = os.path.join(log_dir, output_filename)
        prefill_rows, decode_rows_list = load_prefill_decode(log_dir)
        if not prefill_rows and not any(decode_rows_list):
            print("未解析到任何 Engine 指标行，请确认日志格式。")
            return 0
        _process_one_dir(log_dir, out_path, prefill_rows, decode_rows_list)
        decode_info = ", ".join(f"decode_{i}={len(dr)}" for i, dr in enumerate(decode_rows_list))
        print(f"[单目录] {log_dir}")
        print(f"  Prefill 解析行数: {len(prefill_rows)}, {decode_info}")
        print(f"已保存: {out_path}")
        return 1

    # 当前目录无日志，仅遍历直接子目录（一层）
    print(f"[批量] 输出文件名: {output_filename}")
    batch_list = []
    for name in sorted(os.listdir(log_dir)):
        sub = os.path.join(log_dir, name)
        if os.path.isdir(sub) and _dir_has_logs(sub):
            prefill_rows, decode_rows_list = load_prefill_decode(sub)
            if not prefill_rows and not any(decode_rows_list):
                print(f"  [SKIP] {name} (无有效数据)")
                continue
            batch_list.append((name, prefill_rows, decode_rows_list))
            per_batch_filename = f"{name}_{output_filename}" if not output_filename.startswith(name + "_") else output_filename
            out_path = os.path.join(sub, per_batch_filename)
            _process_one_dir(sub, out_path, prefill_rows, decode_rows_list)
            decode_info = ", ".join(f"d{i}={len(dr)}" for i, dr in enumerate(decode_rows_list))
            print(f"  [OK] {name} -> {per_batch_filename} (prefill={len(prefill_rows)}, {decode_info})")
    if not batch_list:
        print(f"错误：{log_dir} 下既无 prefill.log/decode.log，也无含日志的直接子目录。")
        return 0
    batch_list.sort(key=lambda x: x[0])  # 按子目录名排序
    sweep_filename = "vllm_engine_metrics_sweep.xlsx"
    build_sweep_excel(log_dir, batch_list, sweep_filename)
    print(f"批量完成: {len(batch_list)} 个目录已生成 Excel，并已生成汇总表 {sweep_filename}")
    return len(batch_list)


def main():
    parser = argparse.ArgumentParser(
        description="解析 prefill/decode 日志并生成 Excel 报表与折线图。"
        " 两种用法：1）log_dir 下直接含 prefill.log 与 decode.log（或 decode_0.log、decode_1.log 等）；2）log_dir 下仅有若干子目录，子目录含日志。"
    )
    parser.add_argument(
        "log_dir",
        type=str,
        help="情形1：含 prefill.log/decode.log 的目录；情形2：其直接子目录含上述日志的父目录",
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="输出 Excel 文件名（仅文件名，不含路径）；文件始终生成在 prefill.log 同级目录，默认 vllm_engine_metrics_analysis.xlsx",
    )
    args = parser.parse_args()
    log_dir = os.path.abspath(args.log_dir)
    if not os.path.isdir(log_dir):
        print(f"错误：目录不存在 {log_dir}")
        return 1

    default_filename = "vllm_engine_metrics_analysis.xlsx"
    output_filename = (os.path.basename(args.output).strip() if args.output else default_filename) or default_filename
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    n = process_log_dir(log_dir, output_filename)
    return 0 if n > 0 else 1


if __name__ == "__main__":
    exit(main())
