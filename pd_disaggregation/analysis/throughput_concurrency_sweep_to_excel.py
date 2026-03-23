#!/usr/bin/env python3
"""
从各 batch_* 子目录的 prefill.log / decode.log（或 decode_0.log、decode_1.log、…）中解析「吞吐与并发」指标
（Avg prompt/generation throughput、Running/Waiting reqs、KV cache/prefix cache 等），
生成多并发汇总表 Excel：单 sheet，列名为 BS=N_prefill_*、BS=N_decode_0_*、BS=N_decode_1_* 等
（N 为并发档位；多 decode 时每个 D 单独一列块，不同 decode 视为不同数据）。

日志行示例（vLLM 周期性打印的吞吐与并发状态）：
  Engine 000: Avg prompt throughput: 2304.0 tokens/s, Avg generation throughput: 227.6 tokens/s, Running: 30 reqs, Waiting: 0 reqs, ...

用法：
  python throughput_concurrency_sweep_to_excel.py <log_dir> [--output <文件名.xlsx>]
  # log_dir: 如 log/sharegpt_200_yyz_260313/，其下有 batch_30、batch_40 等子目录，各含 prefill.log、decode.log 或 decode_0/1/...
"""

import re
import os
import glob
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 去掉 ANSI 转义（如 [0;36m(APIServer pid=115079)[0;0m）
ANSI_STRIP = re.compile(r"\x1b\[[0-9;]*m")

# 吞吐与并发指标行：时间戳 + 各数值（Engine 000/001/... 为 vLLM 日志原文，多 decode 时各进程可能为 001、002）
THROUGHPUT_CONCURRENCY_LINE_PATTERN = re.compile(
    r"INFO\s+(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+.*?Engine\s+\d+:\s+"
    r"Avg prompt throughput:\s+([\d.]+)\s+tokens/s,\s+"
    r"Avg generation throughput:\s+([\d.]+)\s+tokens/s,\s+"
    r"Running:\s+(\d+)\s+reqs,\s+Waiting:\s+(\d+)\s+reqs,\s+"
    r"GPU KV cache usage:\s+([\d.]+)%,\s+"
    r"Prefix cache hit rate:\s+([\d.]+)%,\s+"
    r"External prefix cache hit rate:\s+([\d.]+)%"
)

# 指标名（与上面捕获组顺序一致，不含 timestamp）
METRIC_NAMES = [
    "Avg prompt throughput",
    "Avg generation throughput",
    "Running",
    "Waiting",
    "GPU KV cache usage",
    "Prefix cache hit rate",
    "External prefix cache hit rate",
]

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def strip_ansi(line: str) -> str:
    return ANSI_STRIP.sub("", line)


def parse_log_file(log_path: Path) -> list[dict]:
    """解析单个 log 文件，返回按行顺序的字典列表，每项含 timestamp + 各 metric。"""
    rows = []
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            plain = strip_ansi(line)
            m = THROUGHPUT_CONCURRENCY_LINE_PATTERN.search(plain)
            if not m:
                continue
            ts = m.group(1)
            vals = [m.group(i) for i in range(2, 9)]
            row = {"timestamp": ts}
            for name, v in zip(METRIC_NAMES, vals):
                # 数值列：前 5 个为数字，后 3 个为百分比（仍按数字存）
                try:
                    row[name] = float(v)
                except ValueError:
                    row[name] = v
            rows.append(row)
    return rows


def _decode_file_index(path) -> int:
    """从 decode_3.log 提取 3，用于多 Decode 文件排序。"""
    basename = os.path.basename(path)
    m = re.match(r"decode_(\d+)\.log", basename)
    return int(m.group(1)) if m else -1


def collect_batches(log_dir: Path) -> list[str]:
    """收集 log_dir 下所有 batch_* 目录名，按数字排序。"""
    batches = []
    for p in log_dir.iterdir():
        if p.is_dir() and p.name.startswith("batch_"):
            try:
                n = int(p.name.split("_")[1])
                batches.append((n, p.name))
            except (IndexError, ValueError):
                continue
    # 并发数相同（如 batch_60_1P1D_*、batch_60_1P2D_*）时，用目录名做二级排序，保证列顺序稳定
    batches.sort(key=lambda x: (x[0], x[1]))
    return [b[1] for b in batches]


def load_decode_rows_for_batch(batch_path: Path) -> list[list[dict]]:
    """
    加载单个 batch 目录下的 decode 日志。存在 decode.log 则返回 [rows]；
    否则枚举 decode_0.log、decode_1.log、… 按编号排序后返回 [rows_0, rows_1, ...]。
    """
    single = batch_path / "decode.log"
    if single.exists():
        return [parse_log_file(single)]
    pattern = str(batch_path / "decode_*.log")
    files = sorted(glob.glob(pattern), key=_decode_file_index)
    return [parse_log_file(Path(p)) for p in files if _decode_file_index(p) >= 0]


def build_single_sheet(wb, sheet_name: str, prefill_data: dict, decode_data: dict):
    """
    单 sheet：列名为 BS=N_prefill_*、BS=N_decode_0_*、BS=N_decode_1_* 等（多 decode 时每路 D 一块列）。
    先整块 prefill（各 BS 一组列），再按 decode_0、decode_1、… 整块（每块内各 BS）；行按索引对齐，不足留空。
    decode_data[batch] = [decode_0_rows, decode_1_rows, ...]，单 decode 时为 [decode_rows]。
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    def _concurrency_from_batch_name(b: str) -> int:
        try:
            return int(b.split("_")[1])
        except (IndexError, ValueError):
            return -1

    batches = sorted(
        set(prefill_data.keys()) | set(decode_data.keys()),
        key=lambda b: (_concurrency_from_batch_name(b), b),
    )
    if not batches:
        ws["A1"] = "无数据"
        return

    def bs_label(batch: str) -> str:
        # batch_200_1P2D_tokens_8192 -> BS=200_1P2D_tokens_8192
        if batch.startswith("batch_"):
            return f"BS={batch[len('batch_'):]}".replace("__", "_")
        return f"BS={batch}".replace("__", "_")

    cols_per_prefill = 1 + len(METRIC_NAMES)
    cols_per_decode = 1 + len(METRIC_NAMES)

    # 针对每个 batch：decode 列块数量不强制等于全局 max_decodes
    # 这样当某个 batch 实际只有 1 个 decode 时，表头会是 *_decode_timestamp，
    # 而不会出现空的 *_decode_0_timestamp / *_decode_1_timestamp。
    batch_specs = []
    max_rows = 1
    for batch in batches:
        pre_rows = prefill_data.get(batch, [])
        drl = decode_data.get(batch, [])
        dec_cnt = len(drl)
        # 最大行数需要考虑：prefill 与所有 decode 路
        max_rows = max(max_rows, len(pre_rows), max((len(dr) for dr in drl), default=0))
        batch_specs.append((batch, dec_cnt))

    col_idx = 1
    # 表头：每个 batch 先 prefill 块，再按该 batch 自身 dec_cnt 输出 decode 块
    for batch, dec_cnt in batch_specs:
        lb = bs_label(batch)
        ws.cell(1, col_idx, f"{lb}_prefill_timestamp")
        col_idx += 1
        for name in METRIC_NAMES:
            ws.cell(1, col_idx, f"{lb}_prefill_{name}")
            col_idx += 1

        for di in range(dec_cnt):
            # 仅当该 batch 只有 1 路 decode 时才使用不带编号的 decode 标签
            dec_label = "decode" if dec_cnt == 1 else f"decode_{di}"
            ws.cell(1, col_idx, f"{lb}_{dec_label}_timestamp")
            col_idx += 1
            for name in METRIC_NAMES:
                ws.cell(1, col_idx, f"{lb}_{dec_label}_{name}")
                col_idx += 1

    total_cols = col_idx - 1

    for c in range(1, total_cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    for r in range(max_rows):
        col_idx = 1
        for batch, dec_cnt in batch_specs:
            # prefill 块
            rows_list = prefill_data.get(batch, [])
            if r < len(rows_list):
                row = rows_list[r]
                ws.cell(r + 2, col_idx, row.get("timestamp", ""))
                col_idx += 1
                for name in METRIC_NAMES:
                    ws.cell(r + 2, col_idx, row.get(name, ""))
                    col_idx += 1
            else:
                col_idx += cols_per_prefill
            # decode 块（按该 batch 的 dec_cnt 个）
            drl = decode_data.get(batch, [])
            for di in range(dec_cnt):
                if r < len(drl[di]):
                    row = drl[di][r]
                    ws.cell(r + 2, col_idx, row.get("timestamp", ""))
                    col_idx += 1
                    for name in METRIC_NAMES:
                        ws.cell(r + 2, col_idx, row.get(name, ""))
                        col_idx += 1
                else:
                    col_idx += cols_per_decode
        for c in range(1, total_cols + 1):
            ws.cell(r + 2, c).border = BORDER_THIN

    for c in range(1, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20


def main():
    parser = argparse.ArgumentParser(description="从各 batch 的 prefill/decode.log 解析吞吐与并发指标，生成多并发汇总 Excel")
    parser.add_argument("log_dir", type=str, help="包含 batch_30、batch_40 等子目录的路径")
    parser.add_argument("--output", "-o", type=str, default="throughput_concurrency_sweep.xlsx", help="输出 Excel 文件名")
    args = parser.parse_args()

    log_dir = Path(args.log_dir)
    if not log_dir.is_dir():
        raise SystemExit(f"目录不存在: {log_dir}")

    batches = collect_batches(log_dir)
    if not batches:
        raise SystemExit(f"未找到 batch_* 子目录: {log_dir}")

    # 按 log 类型收集：prefill / decode（多 decode 时 decode_data[batch] 为 [decode_0_rows, decode_1_rows, ...]）
    prefill_data = {}
    decode_data = {}

    total_batches = len(batches)
    for bi, batch in enumerate(batches, 1):
        batch_path = log_dir / batch
        prefill_log = batch_path / "prefill.log"

        if prefill_log.exists():
            prefill_data[batch] = parse_log_file(prefill_log)
        else:
            prefill_data[batch] = []

        decode_data[batch] = load_decode_rows_for_batch(batch_path)
        # 防止以为卡住：每遍历一个 batch（子目录）打印一次解析结果概况
        dec_lens = [len(dr) for dr in decode_data.get(batch, [])]
        print(
            f"[throughput_sweep] ({bi}/{total_batches}) batch_dir={batch} "
            f"prefill_rows={len(prefill_data.get(batch, []))} "
            f"decode_files={len(dec_lens)} decode_rows={dec_lens}"
        )

    out_path = log_dir / args.output if not os.path.isabs(args.output) else Path(args.output)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_single_sheet(wb, "throughput_concurrency_sweep", prefill_data, decode_data)

    wb.save(out_path)
    print(f"已写入: {out_path}")

if __name__ == "__main__":
    main()
