"""
merge_and_plot.py

Merges prefix_cache_0313_1.xlsx  and  pd_disaggregation_0314.xlsx into a
single aligned xlsx:
  • Single time(second) column, all series joined on a common time grid.
  • Two-row header: row-1 = group label (merged across metric columns),
                    row-2 = metric names.
  • Charts sheet with all 6 metric plots embedded.

Output: analysis/merge/merged_0313_0314.xlsx
"""

import os, re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.chart.series import SeriesLabel

# ── Paths ─────────────────────────────────────────────────────────────────────
_here   = os.path.dirname(os.path.abspath(__file__))
_ws     = os.path.dirname(_here)
FILE1   = os.path.join(_ws, "analysis", "0313_1",            "prefix_cache_0313_1.xlsx")
FILE2   = os.path.join(_ws, "analysis", "pd_disaggregation", "pd_disaggregation_0314.xlsx")
OUT_DIR = os.path.join(_ws, "analysis", "merge")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Metrics ───────────────────────────────────────────────────────────────────
METRICS = [
    "Prefix cache hit rate",
    "Avg prompt throughput",
    "Avg generation throughput",
    "GPU KV cache usage",
    "Running",
    "Waiting",
]

METRIC_YLABELS = {
    "Prefix cache hit rate":     "Hit Rate (%)",
    "Avg prompt throughput":     "Prompt Throughput (tok/s)",
    "Avg generation throughput": "Gen Throughput (tok/s)",
    "GPU KV cache usage":        "KV Cache Usage (%)",
    "Running":                   "Running Requests",
    "Waiting":                   "Waiting Requests",
}

# ── Parse helpers ─────────────────────────────────────────────────────────────
def _to_elapsed(series):
    """Convert a timestamp column to elapsed seconds, rounded to 10 s grid."""
    ts  = pd.to_datetime(series, format="mixed")
    t0  = ts.iloc[0]
    raw = (ts - t0).dt.total_seconds()
    return (raw / 10).round() * 10   # snap to 10-second grid


def _parse_file1():
    df_raw = pd.read_excel(FILE1)
    bs_set = sorted(set(int(re.search(r"BS=(\d+)", c).group(1))
                        for c in df_raw.columns if re.search(r"BS=(\d+)", c)))
    series = {}
    for bs in bs_set:
        ts_col = f"BS={bs}_timestamp"
        if ts_col not in df_raw.columns:
            continue
        avail = [m for m in METRICS if f"BS={bs}_{m}" in df_raw.columns]
        sub   = df_raw[[ts_col] + [f"BS={bs}_{m}" for m in avail]].dropna(subset=[ts_col]).copy()
        sub["time_s"] = _to_elapsed(sub[ts_col])
        sub.rename(columns={f"BS={bs}_{m}": m for m in avail}, inplace=True)
        sub.drop(columns=[ts_col], inplace=True)
        # keep last value per time tick (de-duplicate)
        series[("0313_1", bs, None)] = (
            sub.groupby("time_s", as_index=False).last()
        )
    return series


def _parse_file2():
    df_raw = pd.read_excel(FILE2, sheet_name="throughput_concurrency_sweep")
    groups = sorted(set(
        (int(re.search(r"BS=(\d+)", c).group(1)),
         re.search(r"BS=\d+_(\w+)_", c).group(1))
        for c in df_raw.columns if re.search(r"BS=\d+_(\w+)_", c)
    ))
    series = {}
    for bs, role in groups:
        ts_col = f"BS={bs}_{role}_timestamp"
        if ts_col not in df_raw.columns:
            continue
        avail = [m for m in METRICS if f"BS={bs}_{role}_{m}" in df_raw.columns]
        sub   = df_raw[[ts_col] + [f"BS={bs}_{role}_{m}" for m in avail]].dropna(subset=[ts_col]).copy()
        sub["time_s"] = _to_elapsed(sub[ts_col])
        sub.rename(columns={f"BS={bs}_{role}_{m}": m for m in avail}, inplace=True)
        sub.drop(columns=[ts_col], inplace=True)
        series[("0314", bs, role)] = (
            sub.groupby("time_s", as_index=False).last()
        )
    return series


def _label(src, bs, role):
    return f"{src}_BS={bs}_{role}" if role else f"{src}_BS={bs}"


# ── Load & align ──────────────────────────────────────────────────────────────
print("Parsing files …")
all_series = {**_parse_file1(), **_parse_file2()}
sorted_keys = sorted(all_series.keys(), key=lambda k: (k[0], k[1], k[2] or ""))

# Common time grid  =  union of all per-series time_s values
all_times = sorted(set().union(*(set(df["time_s"]) for df in all_series.values())))
aligned   = pd.DataFrame({"time(second)": all_times})

# Join every series onto the common grid
for key in sorted_keys:
    src, bs, role = key
    lbl  = _label(src, bs, role)
    df_s = all_series[key].set_index("time_s")
    avail_metrics = [m for m in METRICS if m in df_s.columns]
    df_s = df_s[avail_metrics].rename(columns={m: f"{lbl}|{m}" for m in avail_metrics})
    aligned = aligned.merge(df_s, left_on="time(second)", right_index=True, how="left")

print(f"Aligned table: {len(aligned)} rows × {len(aligned.columns)} columns")

# ── Smoothed Avg prompt throughput columns ────────────────────────────────────
SMOOTH_METRIC = "Avg prompt throughput"
SMOOTH_WINDOW = 10   # 10 ticks × 10 s = 100 s rolling, centre-aligned

smooth_col_names = {}   # key → column name in aligned
sm_extras = {}
for key in sorted_keys:
    lbl = _label(*key)
    src = f"{lbl}|{SMOOTH_METRIC}"
    if src in aligned.columns:
        sm_col = f"{src}|smooth"
        sm_extras[sm_col] = aligned[src].rolling(SMOOTH_WINDOW, min_periods=1, center=True).mean()
        smooth_col_names[key] = sm_col
if sm_extras:
    aligned = pd.concat([aligned, pd.DataFrame(sm_extras, index=aligned.index)], axis=1)

# ── Style constants ───────────────────────────────────────────────────────────
PALETTE = [
    "1F4E79", "2E75B6", "1E6B3A", "6C3483",
    "7D3C98", "1A5276", "0E6655", "784212",
    "922B21", "154360", "145A32", "4A235A",
]
WHITE_FT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FT   = Font(name="Arial", size=9)
TIME_FT   = Font(name="Arial", bold=True, size=9)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center")
THIN      = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)

def _hdr(ws, row, col, value, fill_hex, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = WHITE_FT
    c.fill      = _fill(fill_hex)
    c.alignment = CENTER
    c.border    = THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

# ── Build xlsx ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1 : Merged Data ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Merged Data"
ws.freeze_panes = "B3"

# Row 1: "time(second)" header (spans 2 rows via merge), then group labels
# Row 2: metric names
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 22

# Column 1 – time(second)
ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
_hdr(ws, 1, 1, "time(second)", "1F4E79", width=16)

col_ptr = 2
for ki, key in enumerate(sorted_keys):
    src, bs, role = key
    lbl   = _label(src, bs, role)
    color = PALETTE[ki % len(PALETTE)]

    # find metric columns that belong to this group
    group_metrics = [m for m in METRICS if f"{lbl}|{m}" in aligned.columns]
    n = len(group_metrics)
    if n == 0:
        continue

    # row-1: merged group label
    ws.merge_cells(start_row=1, start_column=col_ptr,
                   end_row=1,   end_column=col_ptr + n - 1)
    _hdr(ws, 1, col_ptr, lbl, color)

    # row-2: individual metric names
    for ci, m in enumerate(group_metrics):
        _hdr(ws, 2, col_ptr + ci, m, color,
             width=max(len(m) + 2, 20))

    col_ptr += n

# Smooth column headers — one group at the end of the header row
smooth_keys = [k for k in sorted_keys if k in smooth_col_names]
if smooth_keys:
    n_sm = len(smooth_keys)
    ws.merge_cells(start_row=1, start_column=col_ptr,
                   end_row=1,   end_column=col_ptr + n_sm - 1)
    _hdr(ws, 1, col_ptr, f"{SMOOTH_METRIC} (smoothed, {SMOOTH_WINDOW * 10}s avg)", "4472C4")
    for ci, key in enumerate(smooth_keys):
        lbl = _label(*key)
        _hdr(ws, 2, col_ptr + ci, lbl, "4472C4", width=max(len(lbl) + 2, 20))

# Data rows — also build col_index_map: "{lbl}|{metric}" -> 1-based column index
col_index_map = {}
col_ptr = 2
for key in sorted_keys:
    lbl           = _label(*key)
    group_metrics = [m for m in METRICS if f"{lbl}|{m}" in aligned.columns]
    for ci, m in enumerate(group_metrics):
        col_index_map[f"{lbl}|{m}"] = col_ptr + ci
    col_ptr += len(group_metrics)

# Smooth column indices — must follow same order as headers above
smooth_col_index_map = {}   # key -> col_idx
for key in smooth_keys:
    smooth_col_index_map[key] = col_ptr
    col_ptr += 1

n_data_rows = len(aligned)

for ri, (_, row_data) in enumerate(aligned.iterrows(), start=3):
    c = ws.cell(row=ri, column=1, value=row_data["time(second)"])
    c.font      = TIME_FT
    c.alignment = CENTER
    c.border    = THIN_BORDER

    for col_key, col_idx in col_index_map.items():
        val = row_data.get(col_key)
        if val is not None and not (isinstance(val, float) and np.isnan(val)):
            c = ws.cell(row=ri, column=col_idx, value=round(float(val), 4))
            c.font      = BODY_FT
            c.alignment = CENTER
            c.border    = THIN_BORDER

    for key, col_idx in smooth_col_index_map.items():
        val = row_data.get(smooth_col_names[key])
        if val is not None and not (isinstance(val, float) and np.isnan(val)):
            c = ws.cell(row=ri, column=col_idx, value=round(float(val), 4))
            c.font      = BODY_FT
            c.alignment = CENTER
            c.border    = THIN_BORDER

# ── Per-BS chart sheets ────────────────────────────────────────────────────────
# Layout (4 rows × 2 cols, 7 charts total):
#   Row 0: [Prefix cache hit rate,         GPU KV cache usage           ]
#   Row 1: [Avg prompt throughput (raw),   Avg prompt throughput (smooth)]
#   Row 2: [Avg generation throughput,     Running                       ]
#   Row 3: [Waiting                                                       ]
CHART_ORDER = [
    # (metric_name, is_smooth)
    ("Prefix cache hit rate",     False),
    ("GPU KV cache usage",        False),
    (SMOOTH_METRIC,               False),   # Avg prompt throughput raw
    (SMOOTH_METRIC,               True),    # Avg prompt throughput smoothed
    ("Avg generation throughput", False),
    ("Running",                   False),
    ("Waiting",                   False),
]

# Width doubled (40 cm); single-column layout to fit.
# ScatterChart gives a true numeric x-axis so time values are correct.
CHART_W   = 40
CHART_H   = 13
SLOT_ROWS = 26       # row spacing per chart (~13 cm / default row height)
LINE_W    = int(1.3 * 12700)   # 1.3 pt in EMUs

# No red, no green — blue/orange/purple/brown/navy/gold
LINE_COLORS = ["1F4E79", "D35400", "6C3483", "784212", "154360", "B8860B"]

all_bs = sorted(set(k[1] for k in sorted_keys))
xvalues = Reference(ws, min_col=1, min_row=3, max_row=2 + n_data_rows)

for bs in all_bs:
    bs_keys = [k for k in sorted_keys if k[1] == bs]
    ws_bs   = wb.create_sheet(f"BS={bs}")

    for midx, (metric, is_smooth) in enumerate(CHART_ORDER):
        chart = ScatterChart()
        chart.title = (f"{metric} (smoothed, {SMOOTH_WINDOW * 10}s avg)"
                       if is_smooth else metric)
        chart.style           = 10
        chart.y_axis.title    = METRIC_YLABELS.get(metric, metric)
        chart.x_axis.title    = "time(second)"
        chart.width           = CHART_W
        chart.height          = CHART_H
        chart.legend.position = "r"
        chart.scatterStyle    = "line"   # lines without markers

        for si, key in enumerate(bs_keys):
            lbl = _label(*key)
            if is_smooth:
                if key not in smooth_col_index_map:
                    continue
                col_idx = smooth_col_index_map[key]
            else:
                col_key = f"{lbl}|{metric}"
                if col_key not in col_index_map:
                    continue
                col_idx = col_index_map[col_key]

            yvalues = Reference(ws, min_col=col_idx,
                                min_row=3, max_row=2 + n_data_rows)
            s = Series(yvalues, xvalues=xvalues)
            s.tx = SeriesLabel(v=lbl)
            s.graphicalProperties.line.width     = LINE_W
            s.graphicalProperties.line.solidFill = LINE_COLORS[si % len(LINE_COLORS)]
            chart.append(s)

        # Single-column layout — charts stacked vertically
        ws_bs.add_chart(chart, f"A{midx * SLOT_ROWS + 1}")

    print(f"  Sheet BS={bs}: {len(CHART_ORDER)} charts")

# ── Save ──────────────────────────────────────────────────────────────────────
out_xlsx = os.path.join(OUT_DIR, "merged_0313_0314.xlsx")
wb.save(out_xlsx)
print(f"Saved xlsx: {out_xlsx}")
print("Done.")
