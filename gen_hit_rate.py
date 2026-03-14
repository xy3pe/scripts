import re
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(description='Generate prefix cache hit rate Excel report.')
parser.add_argument('--task-name', default="0311_vllm_api_stream_chat_multiturn",
                    help='Task name used in log file naming: vllm_{TASK_NAME}_bs{BATCH_SIZE}.log')
parser.add_argument('--log-dir', default=None,
                    help='Directory containing vllm log files (default: ../logs/{TASK_NAME})')
parser.add_argument('--output-dir', default=None,
                    help='Output directory (default: ../analysis/{TASK_NAME})')
args = parser.parse_args()

TASK_NAME  = args.task_name
script_dir = os.path.dirname(os.path.abspath(__file__))
LOG_DIR    = args.log_dir    or os.path.join(script_dir, '..', 'logs',     TASK_NAME)
OUTPUT_DIR = args.output_dir or os.path.join(script_dir, '..', 'analysis', TASK_NAME)
LOG_DIR    = os.path.abspath(LOG_DIR)
OUTPUT_DIR = os.path.abspath(OUTPUT_DIR)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"prefix_cache_{TASK_NAME}.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Parse log files ──────────────────────────────────────────────────────────
log_files = sorted(f for f in os.listdir(LOG_DIR)
                   if f.startswith(f"vllm_{TASK_NAME}_bs") and f.endswith(".log"))

# data: {batch_size: [(timestamp, hit_rate, gen_throughput, kv_cache_usage, prompt_throughput, running, waiting), ...]}
data = {}

for fname in log_files:
    m = re.search(r"_bs(\d+)", fname)
    if not m:
        continue
    batch_size = int(m.group(1))
    entries = []
    with open(os.path.join(LOG_DIR, fname)) as f:
        for line in f:
            if "Prefix cache hit rate:" not in line:
                continue
            ts_m      = re.match(r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", line)
            rate_m    = re.search(r"(?<!External )Prefix cache hit rate: ([0-9.]+)%", line)
            gen_m     = re.search(r"Avg generation throughput: ([0-9.]+) tokens/s", line)
            kv_m      = re.search(r"GPU KV cache usage: ([0-9.]+)%", line)
            prompt_m  = re.search(r"Avg prompt throughput: ([0-9.]+) tokens/s", line)
            running_m = re.search(r"Running: ([0-9]+) reqs", line)
            waiting_m = re.search(r"Waiting: ([0-9]+) reqs", line)
            if ts_m and rate_m:
                entries.append((
                    ts_m.group(1),
                    float(rate_m.group(1)),
                    float(gen_m.group(1))      if gen_m      else None,
                    float(kv_m.group(1))       if kv_m       else None,
                    float(prompt_m.group(1))   if prompt_m   else None,
                    int(running_m.group(1))    if running_m  else None,
                    int(waiting_m.group(1))    if waiting_m  else None,
                ))
    data[batch_size] = entries

batch_sizes = sorted(data.keys())

# ── Styles ───────────────────────────────────────────────────────────────────
HDR_DARK  = PatternFill("solid", start_color="1F4E79")
HDR_MID   = PatternFill("solid", start_color="2E75B6")
HDR_LIGHT = PatternFill("solid", start_color="4472C4")
WHITE_FT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FT   = Font(name="Arial", size=9)
BOLD_FT   = Font(name="Arial", bold=True, size=10)
CENTER    = Alignment(horizontal="center", vertical="center")

def hdr(ws, row, col, value, fill, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = WHITE_FT
    c.fill = fill
    c.alignment = CENTER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────────
# Layout per BS: Timestamp | Hit Rate | Prompt Throughput | Gen Throughput | KV Cache | Running | Waiting
wb = Workbook()
ws_raw = wb.active
ws_raw.title = "Raw Data"
ws_raw.row_dimensions[1].height = 20

col = 1
batch_col_map = {}  # {bs: (ts_col, rate_col, prompt_col, gen_col, kv_col, running_col, waiting_col)}
for bs in batch_sizes:
    hdr(ws_raw, 1, col,   f"BS={bs}  Timestamp",                HDR_DARK,  width=22)
    hdr(ws_raw, 1, col+1, f"BS={bs}  Hit Rate (%)",              HDR_MID,   width=18)
    hdr(ws_raw, 1, col+2, f"BS={bs}  Prompt Throughput (tok/s)", HDR_LIGHT, width=24)
    hdr(ws_raw, 1, col+3, f"BS={bs}  Gen Throughput (tok/s)",    HDR_LIGHT, width=22)
    hdr(ws_raw, 1, col+4, f"BS={bs}  KV Cache (%)",              HDR_MID,   width=18)
    hdr(ws_raw, 1, col+5, f"BS={bs}  Running (reqs)",            HDR_DARK,  width=18)
    hdr(ws_raw, 1, col+6, f"BS={bs}  Waiting (reqs)",            HDR_DARK,  width=18)
    batch_col_map[bs] = (col, col+1, col+2, col+3, col+4, col+5, col+6)
    col += 7

for bs in batch_sizes:
    ts_col, rate_col, prompt_col, gen_col, kv_col, running_col, waiting_col = batch_col_map[bs]
    for r, (ts, rate, gen, kv, prompt, running, waiting) in enumerate(data[bs], start=2):
        ws_raw.cell(row=r, column=ts_col, value=ts).font = BODY_FT
        c_rate = ws_raw.cell(row=r, column=rate_col, value=rate)
        c_rate.font = BODY_FT; c_rate.number_format = "0.0"
        if prompt is not None:
            c_prompt = ws_raw.cell(row=r, column=prompt_col, value=prompt)
            c_prompt.font = BODY_FT; c_prompt.number_format = "0.0"
        if gen is not None:
            c_gen = ws_raw.cell(row=r, column=gen_col, value=gen)
            c_gen.font = BODY_FT; c_gen.number_format = "0.0"
        if kv is not None:
            c_kv = ws_raw.cell(row=r, column=kv_col, value=kv)
            c_kv.font = BODY_FT; c_kv.number_format = "0.0"
        if running is not None:
            c_running = ws_raw.cell(row=r, column=running_col, value=running)
            c_running.font = BODY_FT; c_running.number_format = "0"
        if waiting is not None:
            c_waiting = ws_raw.cell(row=r, column=waiting_col, value=waiting)
            c_waiting.font = BODY_FT; c_waiting.number_format = "0"

# ── Sheet 2: Chart Data ───────────────────────────────────────────────────────
# Three sections side by side, each: Sample# | BS=x col ...
# Section offsets (0-indexed col start within sheet):
#   Hit Rate:       col 1
#   Gen Throughput: col 1 + n+1 + 1  (gap of 1)
#   KV Cache:       col 1 + (n+1+1)*2
n = len(batch_sizes)
SEC_GAP = 1  # blank columns between sections
SEC_W   = 1 + n  # Sample# + n BS columns

def sec_start(sec_idx):
    return 1 + sec_idx * (SEC_W + SEC_GAP)

ws_cd = wb.create_sheet("Chart Data")
ws_cd.row_dimensions[1].height = 20

SECTIONS = [
    ("Sample #", "Hit Rate (%)",             1),   # value index in tuple
    ("Sample #", "Prompt Throughput (tok/s)", 4),
    ("Sample #", "Gen Throughput (tok/s)",   2),
    ("Sample #", "KV Cache (%)",             3),
    ("Sample #", "Running (reqs)",           5),
    ("Sample #", "Waiting (reqs)",           6),
]

for sec_i, (idx_label, metric_label, val_idx) in enumerate(SECTIONS):
    base = sec_start(sec_i)
    # Sample # header
    c = ws_cd.cell(row=1, column=base, value=idx_label)
    c.font = BOLD_FT; c.alignment = CENTER
    ws_cd.column_dimensions[get_column_letter(base)].width = 12
    # BS headers
    for j, bs in enumerate(batch_sizes):
        c = ws_cd.cell(row=1, column=base+1+j, value=f"BS={bs} {metric_label}")
        c.font = BOLD_FT; c.alignment = CENTER
        ws_cd.column_dimensions[get_column_letter(base+1+j)].width = max(len(f"BS={bs} {metric_label}")+2, 14)

max_rows = max(len(data[bs]) for bs in batch_sizes)

for row_i in range(1, max_rows + 1):
    for sec_i, (_, _, val_idx) in enumerate(SECTIONS):
        base = sec_start(sec_i)
        ws_cd.cell(row=row_i+1, column=base, value=row_i).font = BODY_FT
        for j, bs in enumerate(batch_sizes):
            entries = data[bs]
            if row_i <= len(entries):
                v = entries[row_i-1][val_idx]
                if v is not None:
                    c = ws_cd.cell(row=row_i+1, column=base+1+j, value=v)
                    c.font = BODY_FT; c.number_format = "0.0"

# ── Charts ────────────────────────────────────────────────────────────────────
def make_chart(title, y_label, y_min, y_max, num_fmt, sec_i):
    base = sec_start(sec_i)
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = y_label
    chart.x_axis.title = "Sample Index"
    chart.y_axis.numFmt = num_fmt
    chart.y_axis.scaling.min = y_min
    if y_max is not None:
        chart.y_axis.scaling.max = y_max
    chart.height = 16
    chart.width  = 30
    for j, bs in enumerate(batch_sizes):
        n_rows = len(data[bs])
        vals = Reference(ws_cd, min_col=base+1+j, min_row=1, max_row=n_rows+1)
        chart.add_data(vals, titles_from_data=True)
    cats = Reference(ws_cd, min_col=base, min_row=2, max_row=max_rows+1)
    chart.set_categories(cats)
    return chart

chart_hit     = make_chart("Prefix Cache Hit Rate by Batch Size",
                            "Hit Rate (%)",           0, 100, "0.0", 0)
chart_prompt  = make_chart("Avg Prompt Throughput by Batch Size",
                            "Throughput (tokens/s)",  0, None, "0.0", 1)
chart_gen     = make_chart("Avg Generation Throughput by Batch Size",
                            "Throughput (tokens/s)",  0, None, "0.0", 2)
chart_kv      = make_chart("GPU KV Cache Usage by Batch Size",
                            "KV Cache Usage (%)",     0, 100, "0.0", 3)
chart_running = make_chart("Running Requests by Batch Size",
                            "Running (reqs)",         0, None, "0",   4)
chart_waiting = make_chart("Waiting Requests by Batch Size",
                            "Waiting (reqs)",         0, None, "0",   5)

# ── Chart Sheet ───────────────────────────────────────────────────────────────
ws_chart = wb.create_sheet("Chart")
ws_chart.add_chart(chart_hit,     "B2")
ws_chart.add_chart(chart_prompt,  "B35")
ws_chart.add_chart(chart_gen,     "B68")
ws_chart.add_chart(chart_kv,      "B101")
ws_chart.add_chart(chart_running, "B134")
ws_chart.add_chart(chart_waiting, "B167")

wb.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE}")
