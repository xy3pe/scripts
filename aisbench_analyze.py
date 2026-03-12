"""
AISBench Log Analyzer
用途：解析 aisbench 前缀的日志文件，提取统计数据，生成 Excel 报表和折线图。

用法：
    python aisbench_analyze.py --task-name TASK_NAME [--log-dir LOGS_DIR] [--output OUTPUT.xlsx]

日志文件命名规则：aisbench_{TASK_NAME}_bs{BATCH_SIZE}.log
输出文件命名规则：aisbench_{TASK_NAME}.xlsx

默认：
    --log-dir  ../logs/{TASK_NAME}  (相对于脚本所在目录)
    --output   ../analysis/{TASK_NAME}/aisbench_{TASK_NAME}.xlsx
"""

import re
import os
import glob
import subprocess
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


# ── 数据提取 ──────────────────────────────────────────────────────────────────

def extract_stats(filepath):
    """
    从单个 aisbench 日志文件中提取统计数据。
    日志文件包含 ANSI 转义码，使用 strings 命令提取可读文本。
    返回字典，失败返回 None。
    """
    result = subprocess.run(['strings', filepath], capture_output=True, text=True)
    lines = result.stdout.split('\n')
    lines = [l.strip() for l in lines]

    latency_metrics = ['E2EL', 'TTFT', 'TPOT', 'ITL', 'InputTokens', 'OutputTokens', 'OutputTokenThroughput']

    # 找到延迟指标起始位置
    e2el_idx = None
    for i, line in enumerate(lines):
        if line.strip() == 'E2EL':
            e2el_idx = i
            break
    if e2el_idx is None:
        return None

    def parse_value(s):
        m = re.search(r'[\d.]+', s)
        return float(m.group()) if m else None

    data = {}

    # 解析延迟指标（每项格式：name, 'total', avg, min, max, median, p75, p90, p99）
    i = e2el_idx
    for metric in latency_metrics:
        while i < len(lines) and lines[i] != metric:
            i += 1
        if i >= len(lines):
            break
        i += 1
        if i < len(lines) and lines[i] == 'total':
            i += 1
        vals = []
        while i < len(lines) and len(vals) < 8:
            if lines[i] and lines[i] not in latency_metrics and 'Common' not in lines[i]:
                v = parse_value(lines[i])
                if v is not None:
                    vals.append(v)
                    i += 1
                else:
                    break
            else:
                break
        if len(vals) >= 7:
            data[f'{metric}_avg']    = vals[0]
            data[f'{metric}_min']    = vals[1]
            data[f'{metric}_max']    = vals[2]
            data[f'{metric}_median'] = vals[3]
            data[f'{metric}_p75']    = vals[4]
            data[f'{metric}_p90']    = vals[5]
            data[f'{metric}_p99']    = vals[6]

    # 解析 Common Metrics
    common_metrics = {
        'Benchmark Duration':    'benchmark_duration_ms',
        'Total Requests':        'total_requests',
        'Failed Requests':       'failed_requests',
        'Success Requests':      'success_requests',
        'Concurrency':           'concurrency',
        'Max Concurrency':       'max_concurrency',
        'Request Throughput':    'request_throughput_rps',
        'Total Input Tokens':    'total_input_tokens',
        'Prefill Token Throughput': 'prefill_token_throughput_tps',
        'Total Generated Tokens':   'total_generated_tokens',
        'Input Token Throughput':   'input_token_throughput_tps',
        'Output Token Throughput':  'output_token_throughput_tps',
        'Total Token Throughput':   'total_token_throughput_tps',
    }

    cm_idx = None
    for j, line in enumerate(lines):
        if 'Common Metric' in line:
            cm_idx = j
            break

    if cm_idx is not None:
        for metric_name, key in common_metrics.items():
            k = cm_idx
            while k < len(lines) and metric_name not in lines[k]:
                k += 1
            if k < len(lines):
                k += 1
                if k < len(lines) and lines[k] == 'total':
                    k += 1
                if k < len(lines):
                    v = parse_value(lines[k])
                    if v is not None:
                        data[key] = v

    return data


def load_all_logs(log_dir, task_name):
    """扫描目录，加载所有 aisbench_{TASK_NAME}_bs{BATCH_SIZE}.log 文件，按 batch_size 排序返回列表。"""
    pattern = os.path.join(log_dir, f'aisbench_{task_name}_bs*.log')
    files = sorted(glob.glob(pattern))
    all_data = []
    for filepath in files:
        filename = os.path.basename(filepath)
        m = re.search(r'_bs(\d+)\.log', filename)
        batch_size = int(m.group(1)) if m else 0
        stats = extract_stats(filepath)
        if stats:
            stats['file'] = filename
            stats['batch_size'] = batch_size
            all_data.append(stats)
            print(f'  [OK] {filename}: {len(stats)} fields')
        else:
            print(f'  [FAIL] {filename}: no stats found')
    all_data.sort(key=lambda x: x['batch_size'])
    return all_data


# ── Excel 生成 ────────────────────────────────────────────────────────────────

FONT_NAME   = 'Arial'
HDR_FILL    = PatternFill('solid', start_color='1F4E79')
HDR_FONT    = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
SUBHDR_FILL = PatternFill('solid', start_color='2E75B6')
SUBHDR_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
CAT_FILL    = PatternFill('solid', start_color='D6E4F0')
CAT_FONT    = Font(name=FONT_NAME, bold=True, color='1F4E79', size=10)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
CENTER  = Alignment(horizontal='center', vertical='center')
LEFT    = Alignment(horizontal='left',   vertical='center')
R_ALIGN = Alignment(horizontal='right',  vertical='center')
thin    = Side(border_style='thin', color='BFBFBF')
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_raw_data_sheet(ws, all_data, task_name):
    n_conc   = len(all_data)
    DATA_COL = 3  # 数据从 C 列开始

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + n_conc)
    tc = ws.cell(row=1, column=1, value=f'AISBench {task_name} - Benchmark Statistics')
    tc.font = Font(name=FONT_NAME, bold=True, size=14, color='1F4E79')
    tc.alignment = CENTER
    tc.fill = PatternFill('solid', start_color='EBF3FB')

    # 表头行
    for col, val in [(1, 'Metric'), (2, 'Statistic')]:
        c = ws.cell(row=2, column=col, value=val)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER

    for ci, d in enumerate(all_data):
        c = ws.cell(row=2, column=DATA_COL + ci, value=f"BS={d['batch_size']}")
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER

    # 延迟指标分组
    latency_sections = [
        ('E2EL (ms)', [
            ('Average','E2EL_avg'), ('Min','E2EL_min'), ('Max','E2EL_max'),
            ('Median','E2EL_median'), ('P75','E2EL_p75'), ('P90','E2EL_p90'), ('P99','E2EL_p99'),
        ]),
        ('TTFT (ms)', [
            ('Average','TTFT_avg'), ('Min','TTFT_min'), ('Max','TTFT_max'),
            ('Median','TTFT_median'), ('P75','TTFT_p75'), ('P90','TTFT_p90'), ('P99','TTFT_p99'),
        ]),
        ('TPOT (ms)', [
            ('Average','TPOT_avg'), ('Min','TPOT_min'), ('Max','TPOT_max'),
            ('Median','TPOT_median'), ('P75','TPOT_p75'), ('P90','TPOT_p90'), ('P99','TPOT_p99'),
        ]),
        ('ITL (ms)', [
            ('Average','ITL_avg'), ('Min','ITL_min'), ('Max','ITL_max'),
            ('Median','ITL_median'), ('P75','ITL_p75'), ('P90','ITL_p90'), ('P99','ITL_p99'),
        ]),
        ('Output Token\nThroughput (token/s)', [
            ('Average','OutputTokenThroughput_avg'), ('Min','OutputTokenThroughput_min'),
            ('Max','OutputTokenThroughput_max'), ('Median','OutputTokenThroughput_median'),
            ('P75','OutputTokenThroughput_p75'), ('P90','OutputTokenThroughput_p90'),
            ('P99','OutputTokenThroughput_p99'),
        ]),
    ]

    common_section = [
        ('Benchmark Duration (ms)',           'benchmark_duration_ms'),
        ('Total Requests',                    'total_requests'),
        ('Failed Requests',                   'failed_requests'),
        ('Success Requests',                  'success_requests'),
        ('Actual Concurrency',                'concurrency'),
        ('Batch Size (Max Concurrency)',       'max_concurrency'),
        ('Request Throughput (req/s)',         'request_throughput_rps'),
        ('Total Input Tokens',                'total_input_tokens'),
        ('Prefill Token Throughput (token/s)','prefill_token_throughput_tps'),
        ('Total Generated Tokens',            'total_generated_tokens'),
        ('Input Token Throughput (token/s)',  'input_token_throughput_tps'),
        ('Output Token Throughput (token/s)', 'output_token_throughput_tps'),
        ('Total Token Throughput (token/s)',  'total_token_throughput_tps'),
    ]

    row = 3
    for section_name, stats in latency_sections:
        n_rows = len(stats)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + n_rows - 1, end_column=1)
        cat = ws.cell(row=row, column=1, value=section_name)
        cat.font = CAT_FONT; cat.fill = CAT_FILL; cat.border = BORDER
        cat.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for idx, (stat_label, key) in enumerate(stats):
            rf = PatternFill('solid', start_color='F2F8FD') if idx % 2 else None
            sc = ws.cell(row=row, column=2, value=stat_label)
            sc.font = NORMAL_FONT; sc.alignment = LEFT; sc.border = BORDER
            if rf: sc.fill = rf
            for ci, d in enumerate(all_data):
                dc = ws.cell(row=row, column=DATA_COL + ci, value=d.get(key, ''))
                dc.font = NORMAL_FONT; dc.alignment = R_ALIGN
                dc.number_format = '#,##0.00'; dc.border = BORDER
                if rf: dc.fill = rf
            row += 1

    # Common Metrics 分隔标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + n_conc)
    sh = ws.cell(row=row, column=1, value='Common Metrics')
    sh.font = SUBHDR_FONT; sh.fill = SUBHDR_FILL; sh.alignment = CENTER; sh.border = BORDER
    row += 1

    for idx, (label, key) in enumerate(common_section):
        rf = PatternFill('solid', start_color='F2F8FD') if idx % 2 else None
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name=FONT_NAME, bold=True, size=10, color='1F4E79')
        lc.alignment = LEFT; lc.border = BORDER
        if rf: lc.fill = rf
        for ci, d in enumerate(all_data):
            dc = ws.cell(row=row, column=DATA_COL + ci, value=d.get(key, ''))
            dc.font = NORMAL_FONT; dc.alignment = R_ALIGN
            dc.number_format = '#,##0.00'; dc.border = BORDER
            if rf: dc.fill = rf
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    for ci in range(n_conc):
        ws.column_dimensions[get_column_letter(DATA_COL + ci)].width = 18
    ws.freeze_panes = 'C3'
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


def build_charts_sheet(cws, all_data, task_name):
    cws.sheet_view.showGridLines = False

    chart_cols = [
        ('Batch Size',                           'max_concurrency'),
        ('E2EL Avg (ms)',                         'E2EL_avg'),
        ('TTFT Avg (ms)',                         'TTFT_avg'),
        ('TPOT Avg (ms)',                         'TPOT_avg'),
        ('ITL Avg (ms)',                          'ITL_avg'),
        ('Output Token Throughput Avg (token/s)', 'OutputTokenThroughput_avg'),
        ('Request Throughput (req/s)',            'request_throughput_rps'),
        ('Total Token Throughput (token/s)',      'total_token_throughput_tps'),
        ('Input Token Throughput (token/s)',      'input_token_throughput_tps'),
        ('Prefill Token Throughput (token/s)',    'prefill_token_throughput_tps'),
    ]

    # 页面标题
    cws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(chart_cols))
    pt = cws.cell(row=1, column=1, value=f'AISBench Performance Charts - {task_name}')
    pt.font = Font(name=FONT_NAME, bold=True, size=15, color='1F4E79')
    pt.alignment = CENTER

    CD_ROW = 2
    for ci, (hdr, _) in enumerate(chart_cols):
        c = cws.cell(row=CD_ROW, column=1 + ci, value=hdr)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color='FFFFFF')
        c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        cws.column_dimensions[get_column_letter(1 + ci)].width = max(len(hdr) + 2, 14)

    for ri, d in enumerate(all_data):
        r = CD_ROW + 1 + ri
        for ci, (_, key) in enumerate(chart_cols):
            dc = cws.cell(row=r, column=1 + ci, value=d.get(key, 0))
            dc.font = Font(name=FONT_NAME, size=9)
            dc.alignment = R_ALIGN; dc.number_format = '#,##0.00'; dc.border = BORDER

    CD_END_ROW = CD_ROW + len(all_data)

    def make_line_chart(title, y_title, series_col_offsets):
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = y_title
        chart.x_axis.title = 'Batch Size'
        chart.legend.position = 'b'
        chart.y_axis.numFmt = '#,##0.00'
        cats = Reference(cws, min_col=1, min_row=CD_ROW + 1, max_row=CD_END_ROW)
        for col_offset, _ in series_col_offsets:
            data_ref = Reference(cws, min_col=1 + col_offset, min_row=CD_ROW, max_row=CD_END_ROW)
            chart.add_data(data_ref, titles_from_data=True)
            chart.series[-1].smooth = True
        chart.set_categories(cats)
        chart.width = 24
        chart.height = 14
        return chart

    # 图1：延迟
    chart1 = make_line_chart(
        'Latency Metrics (Average) vs Max Concurrency', 'Latency (ms)',
        [(1, 'E2EL'), (2, 'TTFT'), (3, 'TPOT'), (4, 'ITL')]
    )
    cws.add_chart(chart1, 'A11')

    # 图2：Token 吞吐
    chart2 = make_line_chart(
        'Token Throughput vs Max Concurrency', 'Throughput (token/s)',
        [(5, 'Output'), (7, 'Total'), (8, 'Input'), (9, 'Prefill')]
    )
    cws.add_chart(chart2, 'M11')

    # 图3：请求吞吐
    chart3 = make_line_chart(
        'Request Throughput vs Max Concurrency', 'Throughput (req/s)',
        [(6, 'Request Throughput')]
    )
    chart3.width = 24
    chart3.height = 14
    cws.add_chart(chart3, 'A39')


def generate_excel(all_data, output_path, task_name):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Raw Data'
    build_raw_data_sheet(ws, all_data, task_name)

    cws = wb.create_sheet('Charts')
    build_charts_sheet(cws, all_data, task_name)

    wb.save(output_path)
    print(f'Saved: {output_path}')


# ── 入口 ──────────────────────────────────────────────────────────────────────

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    parser = argparse.ArgumentParser(description='Analyze AISBench log files and export to Excel.')
    parser.add_argument('--task-name', required=True,
                        help='Task name used in log file naming: aisbench_{TASK_NAME}_bs{BATCH_SIZE}.log')
    parser.add_argument('--log-dir', default=None,
                        help='Directory containing aisbench log files (default: ../logs/{TASK_NAME})')
    parser.add_argument('--output', default=None,
                        help='Output Excel file path (default: ../analysis/{TASK_NAME}/aisbench_{TASK_NAME}.xlsx)')
    args = parser.parse_args()

    task_name = args.task_name
    log_dir   = os.path.abspath(args.log_dir or os.path.join(script_dir, '..', 'logs', task_name))
    output    = os.path.abspath(args.output  or os.path.join(script_dir, '..', 'analysis', task_name, f'aisbench_{task_name}.xlsx'))

    os.makedirs(os.path.dirname(output), exist_ok=True)

    print(f'Task name     : {task_name}')
    print(f'Log directory : {log_dir}')
    print(f'Output file   : {output}')
    print()

    all_data = load_all_logs(log_dir, task_name)
    if not all_data:
        print('No data found. Check the log directory and task name.')
        return

    print(f'\nLoaded {len(all_data)} log files. Generating Excel...')
    generate_excel(all_data, output, task_name)


if __name__ == '__main__':
    main()
